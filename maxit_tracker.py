# -*- coding: utf-8 -*-
"""
maxit_tracker.py — Suivi Google Play / App Store de l'app Max It (Orange)

Script UNIQUE : toute la config, la logique de scraping et l'écriture Excel
sont dans ce seul fichier. Les deux petits helpers Node.js (un pour Android,
un pour iOS) sont générés automatiquement par ce script au premier lancement
dans un sous-dossier ".node_helpers" - tu n'as donc qu'un seul fichier à
maintenir, mais Node.js reste nécessaire pour exécuter les librairies de
scraping (parse-play et app-store-scraper n'existent qu'en JS).

USAGE
-----
1. (une seule fois) installer les dépendances Node :
       cd .node_helpers && npm install && cd ..
   (le dossier .node_helpers et son package.json sont créés automatiquement
   dès le premier "python maxit_tracker.py", donc lance le script une
   première fois pour les générer, PUIS fais le npm install, PUIS relance.)
2. pip install openpyxl
3. python maxit_tracker.py

Chaque exécution ajoute une ligne par pays x plateforme dans
maxit_store_history.xlsx, sans jamais écraser l'historique précédent.

LIMITES CONNUES
---------------
- Le rang n'est jamais retourné directement par l'API : le script télécharge
  un classement (RANK_SEARCH_DEPTH apps) et cherche Max It dedans. Si
  l'app n'y figure pas, "rank" reste vide (comportement voulu).
- "installs" (nombre de téléchargements) n'existe QUE côté Android. Apple ne
  publie jamais ce chiffre, même approximatif : "installs" reste vide pour iOS.
- Ces deux librairies scrapent des endpoints non documentés des stores : elles
  peuvent casser si Google/Apple changent leur format.
"""

import json
import os
import subprocess
import uuid
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook


# =============================================================================
# 1. CONFIGURATION - pays et identifiants d'app
# =============================================================================
#
# ATTENTION EGYPTE : Max It n'existe pas encore en Egypte. L'app Orange
# Egypte s'appelle toujours "My Orange Egypt" (pas de marque Max it). Les
# identifiants ci-dessous sont corrects pour CETTE app - à toi de décider si
# tu veux la garder dans le suivi ou la retirer de la liste.
#
# category_hint : catégorie utilisée pour chercher le rang de Max It dans un
# classement. "APPLICATION" = classement toutes catégories confondues (comme
# demandé), pas seulement Finance. Le rang trouvé est donc le rang général de
# l'app, tous types d'applications mélangés, pour chaque pays.

COUNTRIES = [
    {"country_code": "GN", "country_name": "Guinée",
     "android_app_id": "com.orange.myorange.ogn", "android_country_ovr": None,
     "ios_app_id": 1177999485, "ios_country_code": "FR",  # Guinée n'a pas de storefront Apple propre
     "category_hint": "APPLICATION"},

    {"country_code": "TN", "country_name": "Tunisie",
     "android_app_id": "com.orange.myorange.otn", "android_country_ovr": None,
     "ios_app_id": 1146892088, "ios_country_code": None,
     "category_hint": "APPLICATION"},

    {"country_code": "SL", "country_name": "Sierra Leone",
     "android_app_id": "com.orange.myorange.osl", "android_country_ovr": None,
     "ios_app_id": 6443607903, "ios_country_code": None,
     "category_hint": "APPLICATION"},

    {"country_code": "SN", "country_name": "Sénégal",
     "android_app_id": "com.orange.myorange.osn", "android_country_ovr": None,
     "ios_app_id": 1039327980, "ios_country_code": None,
     "category_hint": "APPLICATION"},

    {"country_code": "MA", "country_name": "Maroc",
     "android_app_id": "com.orange.meditel.mediteletmoi", "android_country_ovr": None,
     "ios_app_id": 596028698, "ios_country_code": None,
     "category_hint": "APPLICATION"},

    {"country_code": "ML", "country_name": "Mali",
     "android_app_id": "com.oml.dsi.orangemobile", "android_country_ovr": None,
     "ios_app_id": 1494321079, "ios_country_code": None,
     "category_hint": "APPLICATION"},

    {"country_code": "MG", "country_name": "Madagascar",
     "android_app_id": "com.orange.myorange.omg", "android_country_ovr": None,
     "ios_app_id": 1295171817, "ios_country_code": None,
     "category_hint": "APPLICATION"},

    {"country_code": "LR", "country_name": "Liberia",
     "android_app_id": "com.orange.myorange.olr", "android_country_ovr": None,
     "ios_app_id": 6746187327, "ios_country_code": None,
     "category_hint": "APPLICATION"},

    {"country_code": "JO", "country_name": "Jordanie",
     "android_app_id": "com.orange.myorange.ojo", "android_country_ovr": None,
     "ios_app_id": 808824590, "ios_country_code": None,
     "category_hint": "APPLICATION"},

    {"country_code": "CI", "country_name": "Côte d'Ivoire",
     "android_app_id": "com.orange.myorange.oci", "android_country_ovr": None,
     "ios_app_id": 1061120855, "ios_country_code": None,
     "category_hint": "APPLICATION"},

    {"country_code": "GW", "country_name": "Guinée-Bissau",
     "android_app_id": "com.orange.myorange.ogw", "android_country_ovr": None,
     "ios_app_id": 1574668652, "ios_country_code": None,
     "category_hint": "APPLICATION"},

    {"country_code": "CD", "country_name": "RDC",
     "android_app_id": "com.orange.sugu.ocd", "android_country_ovr": None,
     "ios_app_id": 6447896435, "ios_country_code": None,
     "category_hint": "APPLICATION"},

    {"country_code": "CM", "country_name": "Cameroun",
     "android_app_id": "com.orange.myorange.ocm", "android_country_ovr": None,
     "ios_app_id": 1116920093, "ios_country_code": None,
     "category_hint": "APPLICATION"},

    {"country_code": "BF", "country_name": "Burkina Faso",
     "android_app_id": "com.orange.myorange.obf", "android_country_ovr": None,
     "ios_app_id": 1553774707, "ios_country_code": None,
     "category_hint": "APPLICATION"},

    {"country_code": "BW", "country_name": "Botswana",
     "android_app_id": "com.orange.myorange.obw", "android_country_ovr": None,
     "ios_app_id": 1195839458, "ios_country_code": None,
     "category_hint": "APPLICATION"},

    # Egypte : pas encore de Max it, app "My Orange Egypt" classique.
    {"country_code": "EG", "country_name": "Egypte",
     "android_app_id": "com.orange.mobinilandme", "android_country_ovr": None,
     "ios_app_id": 942568333, "ios_country_code": None,
     "category_hint": "APPLICATION"},
]

RANK_SEARCH_DEPTH = 200            # nb d'apps à récupérer pour chercher le rang
OUTPUT_XLSX_PATH = "maxit_store_history.xlsx"
SHEET_NAME = "Historique"
NODE_HELPERS_DIR = ".node_helpers"  # créé automatiquement, contient les .js


# =============================================================================
# 2. CODE DES HELPERS NODE.JS (générés sur disque au premier lancement)
# =============================================================================

PACKAGE_JSON = """{
  "name": "maxit-store-helpers",
  "version": "1.0.0",
  "private": true,
  "dependencies": {
    "parse-play": "^3.1.0",
    "app-store-scraper": "^0.18.0"
  }
}
"""

SCRAPER_ANDROID_JS = r"""
// Généré automatiquement par maxit_tracker.py - basé sur parse-play.
const { fetchAppDetails, fetchTopCharts } = require('parse-play');

async function main() {
  const [appId, country, categoryHint, rankSearchDepthArg] = process.argv.slice(2);
  const rankSearchDepth = parseInt(rankSearchDepthArg, 10) || 200;
  const result = { app_id: appId, country: country };

  try {
    const details = await fetchAppDetails({ appId }, { language: 'EN', country });
    result.app_name = details.name ?? null;
    result.developer = details.developer?.name ?? details.developer ?? null;
    result.category = details.category ?? categoryHint ?? null;
    result.price = details.price ?? 0;
    result.is_free = (details.price ?? 0) === 0;
    result.rating_avg = details.aggregateRating?.ratingValue ?? details.rating ?? null;
    result.rating_count = details.aggregateRating?.ratingCount ?? details.ratingCount ?? null;
    result.installs = details.installs ?? null;
    result.version = details.version ?? null;
    result.description_short = (details.description ?? '').slice(0, 300);
  } catch (err) {
    // DEBUG: on inclut la stack complète pour diagnostiquer (proxy/réseau vs
    // vraie erreur de parsing). A retirer une fois le problème identifié.
    result.error = `fetchAppDetails a échoué: ${err.message}`;
    result.error_stack = err.stack;
    process.stdout.write(JSON.stringify(result));
    return;
  }

  try {
    const topChart = await fetchTopCharts(
      { category: 'APPLICATION', chart: 'topselling_free', count: rankSearchDepth },
      { country, language: 'EN' }
    );
    const idx = topChart.findIndex((app) => app.app_id === appId);
    result.rank = idx === -1 ? null : idx + 1;
  } catch (err) {
    result.rank = null;
    result.rank_error = `fetchTopCharts a échoué: ${err.message}`;
  }

  process.stdout.write(JSON.stringify(result));
}

main().catch((err) => {
  process.stdout.write(JSON.stringify({ error: `Erreur inattendue: ${err.message}` }));
});
"""

SCRAPER_IOS_JS = r"""
// Généré automatiquement par maxit_tracker.py - basé sur app-store-scraper.
const store = require('app-store-scraper');

// app-store-scraper convertit le code pays en identifiant interne de
// vitrine Apple (technologie héritée d'iTunes) pour aller chercher un
// classement. Si le pays n'est pas dans SA table interne, la librairie
// bascule SILENCIEUSEMENT sur le classement des Etats-Unis, sans erreur.
// On récupère cette même table pour détecter ce cas et éviter un résultat
// trompeur (un "None" qui aurait en fait vérifié le mauvais pays).
let knownMarkets = {};
try {
  knownMarkets = require('app-store-scraper/lib/constants').markets;
} catch (e) {
  // Si jamais la structure interne de la librairie change, on continue
  // sans cette vérification plutôt que de faire planter tout le script.
}

async function main() {
  const [appIdArg, country, categoryHint, rankSearchDepthArg] = process.argv.slice(2);
  const rankSearchDepth = parseInt(rankSearchDepthArg, 10) || 200;
  const appIdNum = parseInt(appIdArg, 10);
  const result = { app_id: appIdArg, country: country };

  try {
    const details = await store.app({ id: appIdNum, country });
    result.app_name = details.title ?? null;
    result.developer = details.developer ?? null;
    result.category = details.primaryGenre ?? categoryHint ?? null;
    result.price = details.price ?? 0;
    result.is_free = details.free ?? (details.price === 0);
    result.rating_avg = details.score ?? null;
    result.rating_count = details.reviews ?? null;
    result.installs = null; // Apple ne publie jamais ce chiffre
    result.version = details.version ?? null;
    result.description_short = (details.description ?? '').slice(0, 300);
  } catch (err) {
    result.error = `store.app a échoué: ${err.message}`;
    process.stdout.write(JSON.stringify(result));
    return;
  }

  const countryIsSupportedForRanking = knownMarkets[country.toUpperCase()] !== undefined;

  if (!countryIsSupportedForRanking) {
    // app-store-scraper ne sait pas interroger ce pays (table interne
    // obsolète), mais Apple a bien une vraie vitrine pour ce pays - on
    // scrape donc directement la page web publique des classements.
    // Limite : cette page ne montre que le TOP 25 (contre top 200 pour
    // les pays supportés nativement par la librairie), donc si Max it
    // est classé au-dela de la 25e place, "rank" restera vide ici même
    // si un rang reel existe plus loin.
    try {
      const chartUrl = `https://apps.apple.com/${country.toLowerCase()}/iphone/charts/36?chart=top-free`;
      const res = await fetch(chartUrl, {
        headers: { 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)' },
      });
      const html = await res.text();

      // Extrait les ID d'app dans leur ordre d'apparition dans la page
      // (qui correspond a l'ordre du classement).
      const idMatches = [...html.matchAll(/\/id(\d+)"/g)].map((m) => m[1]);
      const idsUniquesOrdonnes = [...new Set(idMatches)];

      const idx = idsUniquesOrdonnes.findIndex((id) => Number(id) === appIdNum);
      if (idx === -1) {
        result.rank = null;
        result.rank_note = `Non trouvé dans le top 25 web (méthode de repli limitée à 25, pays "${country}" non couvert par app-store-scraper).`;
      } else {
        result.rank = idx + 1;
        result.rank_note = 'Rang obtenu via la page web des classements (méthode de repli, top 25 max), pas via app-store-scraper.';
      }
    } catch (err) {
      result.rank = null;
      result.rank_note = `Échec du scraping de repli pour "${country}": ${err.message}`;
    }
    process.stdout.write(JSON.stringify(result));
    return;
  }

  try {
    const topList = await store.list({
      collection: store.collection.TOP_FREE_IOS,
      // Pas de filtre "category" ici : on veut le classement toutes
      // catégories confondues (comme sur Android avec 'APPLICATION'),
      // pas seulement la catégorie Finance.
      country,
      num: rankSearchDepth,
    });
    const idx = topList.findIndex((app) => Number(app.id) === appIdNum);
    result.rank = idx === -1 ? null : idx + 1;
  } catch (err) {
    result.rank = null;
    result.rank_error = `store.list a échoué: ${err.message}`;
  }

  process.stdout.write(JSON.stringify(result));
}

main().catch((err) => {
  process.stdout.write(JSON.stringify({ error: `Erreur inattendue: ${err.message}` }));
});
"""


def ensure_node_helpers_exist():
    """Écrit les fichiers Node (package.json + les 2 scrapers) sur disque
    s'ils n'existent pas encore. Ne les écrase pas s'ils existent déjà, pour
    ne pas perdre un éventuel npm install / node_modules existant."""
    os.makedirs(NODE_HELPERS_DIR, exist_ok=True)

    files = {
        "package.json": PACKAGE_JSON,
        "scraper_android.js": SCRAPER_ANDROID_JS,
        "scraper_ios.js": SCRAPER_IOS_JS,
    }
    for filename, content in files.items():
        path = os.path.join(NODE_HELPERS_DIR, filename)
        if not os.path.exists(path):
            with open(path, "w", encoding="utf-8") as f:
                f.write(content)


# =============================================================================
# 3. APPEL DES HELPERS NODE + CONSTRUCTION DES LIGNES
# =============================================================================

COLUMNS = [
    "run_id", "run_ts_utc", "snapshot_date",
    "country_code", "country_name", "platform", "app_id",
    "app_name", "developer", "category", "price", "is_free",
    "rating_avg", "rating_count", "rank", "rank_note", "installs", "version",
    "description_short", "status", "error_message",
]


def run_node_helper(script_name, app_id, country, category_hint):
    script_path = os.path.join(NODE_HELPERS_DIR, script_name)
    args = ["node", script_path, str(app_id), str(country),
            str(category_hint or ""), str(RANK_SEARCH_DEPTH)]

    completed = subprocess.run(
        args, capture_output=True, text=True, timeout=60,
        encoding="utf-8", errors="replace",
    )

    if completed.returncode != 0:
        raise RuntimeError(
            f"Le process Node a retourné le code {completed.returncode}. "
            f"stderr: {completed.stderr.strip()}"
        )

    stdout = completed.stdout.strip()
    if not stdout:
        raise RuntimeError("Le process Node n'a rien renvoyé sur stdout.")

    return json.loads(stdout)


def build_row(run_id, run_ts_utc, snapshot_date, country_cfg, platform):
    country_code = country_cfg["country_code"]
    country_name = country_cfg.get("country_name", "")
    category_hint = country_cfg.get("category_hint")

    base_row = {
        "run_id": run_id, "run_ts_utc": run_ts_utc, "snapshot_date": snapshot_date,
        "country_code": country_code, "country_name": country_name, "platform": platform,
    }

    if platform == "android":
        app_id = country_cfg.get("android_app_id")
        store_country = country_cfg.get("android_country_ovr") or country_code
        script_name = "scraper_android.js"
    elif platform == "ios":
        app_id = country_cfg.get("ios_app_id")
        store_country = country_cfg.get("ios_country_code") or country_code
        script_name = "scraper_ios.js"
    else:
        raise ValueError(f"Plateforme inconnue: {platform}")

    if not app_id:
        return None  # pas configuré pour ce pays -> ignoré, pas une erreur

    base_row["app_id"] = app_id

    try:
        data = run_node_helper(script_name, app_id, store_country, category_hint)

        if "error" in data:
            base_row["status"] = "error"
            # DEBUG: on inclut la stack Node si présente, pour diagnostiquer.
            msg = data["error"]
            if data.get("error_stack"):
                msg += " | STACK: " + data["error_stack"].replace("\n", " ")
            base_row["error_message"] = msg
            return base_row

        base_row.update({
            "app_name": data.get("app_name"),
            "developer": data.get("developer"),
            "category": data.get("category"),
            "price": data.get("price"),
            "is_free": data.get("is_free"),
            "rating_avg": data.get("rating_avg"),
            "rating_count": data.get("rating_count"),
            "rank": data.get("rank"),
            "rank_note": data.get("rank_note") or data.get("rank_error", ""),
            "installs": data.get("installs"),
            "version": data.get("version"),
            "description_short": data.get("description_short"),
            "status": "ok",
            "error_message": "",
        })
        return base_row

    except Exception as exc:
        base_row["status"] = "error"
        base_row["error_message"] = str(exc)
        return base_row


def collect_all_rows():
    run_id = str(uuid.uuid4())
    now_utc = datetime.now(timezone.utc)
    run_ts_utc = now_utc.strftime("%Y-%m-%d %H:%M:%S")
    snapshot_date = now_utc.strftime("%Y-%m-%d")

    rows = []
    summary = {"ok": 0, "error": 0, "skipped": 0}

    for country_cfg in COUNTRIES:
        for platform in ("android", "ios"):
            print(f"→ {country_cfg['country_code']} / {platform} ...")
            row = build_row(run_id, run_ts_utc, snapshot_date, country_cfg, platform)

            if row is None:
                summary["skipped"] += 1
                print(f"  (ignoré : pas d'app_id {platform} configuré)")
                continue

            rows.append(row)

            if row["status"] == "ok":
                summary["ok"] += 1
                print(f"  OK - rang: {row.get('rank')}, note: {row.get('rating_avg')}")
            else:
                summary["error"] += 1
                print(f"  ERREUR - {row['error_message']}")

    return rows, summary


# =============================================================================
# 4. ECRITURE EXCEL (créer si absent / ajouter si existant)
# =============================================================================

def append_rows_to_excel(xlsx_path, sheet_name, rows):
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(COLUMNS)
        else:
            ws = wb[sheet_name]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(COLUMNS)

    for row_dict in rows:
        ws.append([row_dict.get(col, "") for col in COLUMNS])

    wb.save(xlsx_path)


# =============================================================================
# 5. POINT D'ENTREE
# =============================================================================

def main():
    print("=== Max It Store Tracker - démarrage ===")
    ensure_node_helpers_exist()
    print(f"Fichier de sortie : {OUTPUT_XLSX_PATH}")
    print(f"Nombre de pays configurés : {len(COUNTRIES)}")
    print()

    rows, summary = collect_all_rows()

    if rows:
        append_rows_to_excel(OUTPUT_XLSX_PATH, SHEET_NAME, rows)
        print(f"\n{len(rows)} ligne(s) ajoutée(s) dans {OUTPUT_XLSX_PATH}")
    else:
        print("\nAucune ligne à écrire (vérifie COUNTRIES dans le script).")

    print("\n=== Résumé ===")
    print(f"  Succès : {summary['ok']}")
    print(f"  Erreurs : {summary['error']}")
    print(f"  Ignorés (app_id non configuré) : {summary['skipped']}")
    print("=== Terminé ===")


if __name__ == "__main__":
    main()